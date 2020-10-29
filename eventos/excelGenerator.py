import openpyxl
from config.settings import BASE_DIR

def genExcel(invitados):

    doc = openpyxl.load_workbook(BASE_DIR + '/static/' + 'sheet.xlsx')

    sheets = doc.sheetnames #.sheetnames accede a las hojas dentro del doc y las convierte en lista

    hoja1 = doc['Hoja 1']#Guardamos en una variable la info de una hoja

    hoja1['A1'] = "Nombre"
    hoja1['B1'] = "Numero"
    hoja1['C1'] = "Mesa"

    for i in range(len(invitados)):
        hoja1['A' + str(i+2)] = invitados[i].nombre
        hoja1['B' + str(i+2)] = invitados[i].numero
        if invitados[i].mesa is not None:
            hoja1['C' + str(i + 2)] = invitados[i].mesa.numero_mesa
        else:
            hoja1['C' + str(i + 2)] = "N/A"

    doc.save(BASE_DIR + '/static/' + 'InvitadosEvento.xlsx')